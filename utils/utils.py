from copy import deepcopy
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from datetime import datetime
from openpyxl import load_workbook
from icecream import ic

def _apply_offset(program: list[list], offset: float) -> list[list]:
    cycle_time = sum([time_duration for _, time_duration in program])
    new_program = deepcopy(program)
    accum = offset % cycle_time
    for phase, duration in reversed(program):
        if accum >= duration:
            new_program[:0] = [[phase, duration]]
            new_program.pop()
            accum -= duration
            if accum == 0: break
        elif accum < duration: # accum < duration
            new_program[:0] = [[phase, accum]]
            new_program[-1][1] = duration - accum
            accum = 0
            break

    return new_program

def _get_color_values(lead_lag, green_val, blue_val):
    if lead_lag == "LEAD":
        return[["blue", blue_val], ["green", green_val]]
    elif lead_lag == "LAG":
        return[["green", green_val], ["blue", blue_val]]
    else: # Caso None
        return [["green", green_val], ["blue", 0]]

def _draw_light(ax, position, cycle, time_max, program, color_green="green", color_left="blue", color_amber="yellow", color_red="red"):
    for t in np.arange(0, time_max, cycle):
        # Dibujar semáforo inbound
        _draw_segments(
            ax,
            [position, position + 10],
            t,
            program["inbound"],
        )
        # Dibujar semáforo outbound
        _draw_segments(
            ax,
            [position - 10, position],
            t,
            program["outbound"],
        )

def _draw_segments(ax, position_range, t_start, phases: list[list]) -> None:
    t = t_start
    for colour, duration in phases:
        ax.fill_betweenx(
            position_range,
            t,
            t + duration,
            color=colour,
            alpha=1,
        )
        t += duration

def gps_tracking(gps_path: str, inbound: bool, distances: list[int], threshold_hour: int, displacement: int, last_offset: int=0) -> list[pd.DataFrame, pd.DataFrame]:

    with open(gps_path, "r", encoding="latin-1") as file:
        lines = file.readlines()

    data = []
    start_parsing = False

    for line in lines:
        line = line.strip()

        if line.startswith("Header"):
            start_parsing = True
            continue

        if start_parsing and line.startswith("Trackpoint"):
            values = line.split("\t")

            while len(values )< 10:
                values.append("")

            # Extraer los valores con la estructura esperada
            hour_string = datetime.strptime(values[2].strip(), "%d/%m/%Y %I:%M:%S %p")
            time = hour_string.strftime("%H:%M:%S")
            leg_length = float(values[6][:-2]) if values[6] else 0.0
            leg_speed = float(values[8][:-4]) if values[8] else 0.0

            # Agregar a la lista de datos
            data.append([time, leg_length, leg_speed])

    # Crear el DataFrame
    columns = ["Time", "Leg Length", "Leg Speed"]
    df_gps = pd.DataFrame(data, columns=columns)

    # Acumulando distancias
    df_gps["Cumulative Length"] = df_gps["Leg Length"].cumsum()

    # Calculando upper_limit
    df_gps["Time"] = pd.to_datetime(df_gps["Time"], format="%H:%M:%S")

    # Conviertiendo el upper_hour a datetime
    threshold_hour = pd.to_datetime(threshold_hour, format="%H:%M:%S")

    df_gps = df_gps[df_gps["Time"] >= threshold_hour]

    df_gps["Leg Length"] = pd.to_numeric(df_gps["Leg Length"], errors="coerce")
    df_gps["Leg Speed"] = pd.to_numeric(df_gps["Leg Speed"], errors="coerce")

    if inbound:
        df_gps["Distance"] = sum(distances) - (df_gps["Cumulative Length"] - df_gps["Cumulative Length"].iloc[0])
    else:
        df_gps["Distance"] = df_gps["Cumulative Length"] - df_gps["Cumulative Length"].iloc[0]

    df_gps["Time_Seconds"] = (df_gps["Time"] - df_gps["Time"].iloc[0]).dt.total_seconds()

    df_gps["Time_Seconds"] += displacement + last_offset

    return df_gps

def read_program(excel_path: str) -> tuple[dict, list, list]:
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active #NOTE: Podría generar un error

    n = 0 #NOTE: Uso para saltos de filas
    programs = {}
    offsets = []
    distances = [0]
    speeds = {
        "inbound": [],
        "outbound": [],
    }
    while True:
        name = ws.cell(row=3+2*n, column=1).value
        if name == None:
            break

        # Guardar más información
        id = n+1
        red =               ws.cell(row=3+2*n, column=12).value

        inbound_green =     ws.cell(row=3+2*n, column=6).value
        inbound_blue =      ws.cell(row=3+2*n, column=10).value
        inbound_leadlag =   ws.cell(row=3+2*n, column=8).value

        outbound_green =    ws.cell(row=3+2*n, column=9).value
        outbound_blue =     ws.cell(row=3+2*n, column=7).value
        outbound_leadlag =  ws.cell(row=3+2*n, column=11).value

        # Construcción del programa dinámico
        program = {
            "inbound": _get_color_values(inbound_leadlag, inbound_green, inbound_blue) + [["yellow", 3], ["red", red]],
            "outbound": _get_color_values(outbound_leadlag, outbound_green, outbound_blue) + [["yellow", 3], ["red", red]],
        }
        
        # Desfases
        offset = ws.cell(row=3+2*n, column=5).value
        offsets.append(offset)

        # Distancias
        distance = ws.cell(row=4+2*n, column=2).value
        if distance != None:
            distances.append(distance)

        # Velocidades
        in_speed = ws.cell(row=4+2*n, column=3).value
        if in_speed != None:
            in_speed_ms = in_speed / 3.6
            speeds["inbound"].append(in_speed_ms)
        out_speed = ws.cell(row=4+2*n, column=4).value
        if out_speed != None:
            out_speed_ms = out_speed / 3.6
            speeds["outbound"].append(out_speed_ms)

        # Programación
        programs[id] = program
        n += 1

    wb.close()
    
    return programs, offsets, distances, speeds

def draw_tracking(
    original_programs: dict,
    offsets: list,
    distances: list,
    speeds: dict,
    dfs_inbound: list = [],
    dfs_outbound: list = [],
    number_cycles: int = 7
) -> None:
    intersection_positions = [sum(distances)-sum(distances[:i+1]) for i in range(len(distances))]
    programs = deepcopy(original_programs)

    for i, (intersection, dict_bounds) in enumerate(original_programs.items()):
        for bound, program_list in dict_bounds.items():
            new_program = _apply_offset(program_list, offsets[i])
            programs[intersection][bound] = new_program

    light_cycles = {number: sum([value for color_key, value in dict_bounds["inbound"] if color_key != "yellow"]) for number, dict_bounds in programs.items()}
    time_max = max(light_cycles.values())*number_cycles

    # Dibujo de bandas de ola verde

    fig, ax = plt.subplots(figsize=(10, 6))
    ax.set_xlim(0, time_max)
    ax.set_xlabel("Time (s)")
    ax.set_ylabel("Distance (m)")
    ax.set_title("Time-Space Diagram")

    # draw_band(ax, intersection_positions, programs, speeds, offsets, time_max, light_cycles, direction="inbound", color="green")
    # draw_band(ax, intersection_positions, programs, speeds, offsets, time_max, light_cycles, direction="outbound", color="blue")

    ax.legend()

    # Dibujo de tracking de GPS

    if dfs_inbound is not None: #TODO: Agregar el nombre del archivo :D
        for df_in in dfs_inbound:
            ax.plot(df_in["Time_Seconds"], df_in["Distance"], marker='o', linestyle='-', color='purple', label='Recorrido de Ida')

    if dfs_outbound is not None:
        for df_out in dfs_outbound:
            ax.plot(df_out["Time_Seconds"], df_out["Distance"], marker='o', linestyle='-', color='cyan', label='Recorrido de Vuelta')

    ax.legend()

    # Dibujar semáforos
    for (i, position), program, cycle in zip(enumerate(intersection_positions), programs.values(), light_cycles.values()):
        _draw_light(ax, position, cycle, time_max, program)

    margin = 30
    ax.set_ylim(min(intersection_positions)-margin, max(intersection_positions)+margin)

    ax2 = ax.twinx()
    ax2.set_ylim(ax.get_ylim())
    ax2.set_yticks(intersection_positions)
    ax2.set_yticklabels(range(1, len(intersection_positions)+1))
    ax2.set_ylabel("Intersecciones")

    # Ajustes finales
    ax.grid(True, linestyle="--", alpha=0.7)
    plt.tight_layout()
    plt.show()