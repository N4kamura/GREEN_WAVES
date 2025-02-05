from copy import deepcopy
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from datetime import datetime
from openpyxl import load_workbook
from icecream import ic

def apply_offset(program: list[list], offset: float) -> list[list]:
    cycle_time = sum([time_duration for _, time_duration in program])
    new_program = deepcopy(program)
    accum = offset % cycle_time
    for phase, duration in reversed(program):
        if accum >= duration:
            new_program[:0] = [[phase, duration]]
            new_program.pop()
            accum -= duration
            if accum == 0: break
        elif accum < duration: # accum < duration
            new_program[:0] = [[phase, accum]]
            new_program[-1][1] = duration - accum
            accum = 0
            break

    return new_program

def _get_color_values(lead_lag, green_val, blue_val):
    if lead_lag == "LEAD":
        return[["blue", blue_val], ["green", green_val]]
    elif lead_lag == "LAG":
        return[["green", green_val], ["blue", blue_val]]
    else: # Caso None
        return [["green", green_val], ["blue", 0]]

def draw_light(ax, position, cycle, time_max, program, color_green="green", color_left="blue", color_amber="yellow", color_red="red"):
    for t in np.arange(0, time_max, cycle):
        # Dibujar semáforo inbound
        _draw_segments(
            ax,
            [position, position + 10],
            t,
            program["inbound"],
        )
        # Dibujar semáforo outbound
        _draw_segments(
            ax,
            [position - 10, position],
            t,
            program["outbound"],
        )

def _draw_segments(ax, position_range, t_start, phases: list[list]) -> None:
    t = t_start
    for colour, duration in phases:
        ax.fill_betweenx(
            position_range,
            t,
            t + duration,
            color=colour,
            alpha=1,
        )
        t += duration

def draw_band(ax, intersection_positions: list, programs: dict, speeds: dict, offsets: list, time_max, light_cycles, direction="inbound", color="green"):
    """
    Dibuja las bandas de tiempo-espacio para un flujo específico (inbound u outbound).
    
    Args:
        ax: Ejes del gráfico.
        intersection_positions: Lista de posiciones de las intersecciones.
        programs: Diccionario de programas semafóricos.
        velocity: Velocidad en m/s.
        time_max: Duración máxima del gráfico (en segundos).
        direction: Dirección del flujo ('inbound' o 'outbound').
        color: Color de la banda.
    """
    offsets_out = offsets[1:] # No se considera el primero que es 0

    offsets_in = offsets[:-1]
    offsets_in = offsets_in[::-1]
    aux_programs_in = programs.copy()
    aux_programs_in.popitem()
    programs_in = dict(reversed(aux_programs_in.items()))

    programs_out = programs.copy()
    first_key = next(iter(programs_out))
    programs_out.pop(first_key)

    if direction == "inbound":
        positions = intersection_positions[::-1]
        in_speeds = speeds["inbound"]
        for i, (start_pos, end_pos) in enumerate(zip(positions[1:], positions[:-1])):
            slope = -1/in_speeds[i]
            cycle = light_cycles[i + 1]  # Ciclo total de la intersección
            list_program = programs_in[len(programs_in)-i][direction]
            green_duration = 0
            for colour, time in list_program:
                if colour == "green":
                    green_duration += time # Duración del verde
            start_time = 0 + offsets_in[i]  # Inicia en t=0

            while start_time < time_max:
                # Coordenadas del inicio y fin de la banda verde
                x1 = start_time
                y1 = start_pos
                x2 = x1 + slope * (end_pos - start_pos)  # Avance en el eje tiempo
                y2 = end_pos

                # Coordenadas del final de la banda verde
                x1_end = x1 + green_duration
                x2_end = x2 + green_duration

                # Dibujar la banda verde
                ax.fill(
                    [x1, x2, x2_end, x1_end],
                    [y1, y2, y2, y1],
                    color=color,
                    alpha=0.2
                )

                # Avanzar al siguiente ciclo
                start_time += cycle

    elif direction == "outbound":
        # Velocidad en m/s
        # slope = 1 / velocity  # Pendiente para el eje tiempo-espacio
        positions = intersection_positions
        out_speeds = speeds["outbound"]
        for i, (start_pos, end_pos) in enumerate(zip(positions[1:], positions[:-1])):
            slope = 1 / out_speeds[i]
            cycle = light_cycles[i + 2]  # Ciclo total de la intersección
            list_program = programs_out[i+2][direction] #TODO
            green_duration = 0
            for colour, time in list_program:
                if colour == "green":
                    green_duration += time # Duración del verde
            start_time = 0 + offsets_out[i] # Inicia en t=0

            while start_time < time_max:
                # Coordenadas del inicio y fin de la banda verde
                x1 = start_time
                y1 = start_pos
                x2 = x1 + slope * (end_pos - start_pos)  # Avance en el eje tiempo
                y2 = end_pos

                # Coordenadas del final de la banda verde
                x1_end = x1 + green_duration
                x2_end = x2 + green_duration

                # Dibujar la banda verde
                ax.fill(
                    [x1, x2, x2_end, x1_end],
                    [y1, y2, y2, y1],
                    color=color,
                    alpha=0.2
                )

                # Avanzar al siguiente ciclo
                start_time += cycle

def read_gps_data(in_path: str, out_path: str, inbound: bool, outbound: bool, lower_hour_out: str, lower_hour_in: str, distances: list[int], last_offset: int=0, displacement_in: int=0, displacement_out: int=0) -> list[pd.DataFrame, pd.DataFrame]:
    ###########
    # Inbound #
    ###########

    if inbound:
        with open(in_path, "r") as file:
            lines = file.readlines()

        # Lista para almacenar los datos procesados
        data = []

        # Identificar la sección de datos (evitar encabezados)
        start_parsing = False

        for line in lines:
            line = line.strip()  # Eliminar espacios en blanco y saltos de línea
            
            # Detectar el inicio de los datos cuando aparece "Trackpoint"
            if line.startswith("Header"):
                start_parsing = True
                continue  # Saltamos la línea de encabezado
            
            if start_parsing and line.startswith("Trackpoint"):
                # Separar por tabulaciones
                values = line.split("\t")
                
                # Manejo de datos vacíos (considerando que Temperature está vacío en algunos casos)
                while len(values) < 10:  # Ajustar longitud de lista si hay columnas vacías
                    values.append("")

                # Extraer los valores con la estructura esperada
                hour_string = datetime.strptime(values[2].strip(), "%d/%m/%Y %I:%M:%S %p")
                time = hour_string.strftime("%H:%M:%S")
                leg_length = float(values[6][:-2]) if values[6] else 0.0
                leg_speed = float(values[8][:-4]) if values[8] else 0.0

                # Agregar a la lista de datos
                data.append([time, leg_length, leg_speed])

        # Crear el DataFrame
        columns = ["Time", "Leg Length", "Leg Speed"]
        df_in = pd.DataFrame(data, columns=columns)

        # Acumulando distancias
        df_in["Cumulative Length"] = df_in["Leg Length"].cumsum()

        # Calculando upper_limit
        df_in["Time"] = pd.to_datetime(df_in["Time"], format="%H:%M:%S")

        # Conviertiendo el upper_hour a datetime
        lower_time = pd.to_datetime(lower_hour_in, format="%H:%M:%S")

        # Excluyendo los datos posteriores al upper_time
        df_in = df_in[df_in["Time"] >= lower_time]

        # Convertir valores numéricos
        df_in["Leg Length"] = pd.to_numeric(df_in["Leg Length"], errors="coerce")
        df_in["Leg Speed"] = pd.to_numeric(df_in["Leg Speed"], errors="coerce")

        # Normalizando a 0:
        df_in["Distance"] = sum(distances) - (df_in["Cumulative Length"] - df_in["Cumulative Length"].iloc[0])
        # df_in["Distance"] = df_in["Cumulative Length"] - lower_limit

        # Convertir la columna 'Time' a datetime
        df_in["Time_Seconds"] = (df_in["Time"] - df_in["Time"].iloc[0]).dt.total_seconds()

        # Aplicar demora para afinar los inicios de los tiempos
        df_in["Time_Seconds"] += displacement_in #NOTE: Por si paso el medio de la intersección unos segundos después

    ############
    # Outbound #
    ############

    if outbound:
        with open(out_path, "r") as file:
            lines = file.readlines()

        # Lista para almacenar los datos procesados
        data = []

        # Identificar la sección de datos (evitar encabezados)
        start_parsing = False

        for line in lines:
            line = line.strip()  # Eliminar espacios en blanco y saltos de línea
            
            # Detectar el inicio de los datos cuando aparece "Trackpoint"
            if line.startswith("Header"):
                start_parsing = True
                continue  # Saltamos la línea de encabezado
            
            if start_parsing and line.startswith("Trackpoint"):
                # Separar por tabulaciones
                values = line.split("\t")
                
                # Manejo de datos vacíos (considerando que Temperature está vacío en algunos casos)
                while len(values) < 10:  # Ajustar longitud de lista si hay columnas vacías
                    values.append("")

                # Extraer los valores con la estructura esperada
                hour_string = datetime.strptime(values[2].strip(), "%d/%m/%Y %I:%M:%S %p")
                time = hour_string.strftime("%H:%M:%S")
                leg_length = float(values[6][:-2]) if values[6] else 0.0
                leg_speed = float(values[8][:-4]) if values[8] else 0.0

                # Agregar a la lista de datos
                data.append([time, leg_length, leg_speed])

        # Crear el DataFrame
        columns = ["Time", "Leg Length", "Leg Speed"]
        df_out = pd.DataFrame(data, columns=columns)

        # Acumulando distancias
        df_out["Cumulative Length"] = df_out["Leg Length"].cumsum()

        # Calculando upper_limit
        df_out["Time"] = pd.to_datetime(df_out["Time"], format="%H:%M:%S")

        # Conviertiendo el upper_hour a datetime
        lower_time = pd.to_datetime(lower_hour_out, format="%H:%M:%S")

        # Excluyendo los datos posteriores al upper_time
        df_out = df_out[df_out["Time"] >= lower_time]

        # Convertir valores numéricos
        df_out["Leg Length"] = pd.to_numeric(df_out["Leg Length"], errors="coerce")
        df_out["Leg Speed"] = pd.to_numeric(df_out["Leg Speed"], errors="coerce")

        # Si no es inbound, calcular "Decreasing Length"
        df_out["Distance"] = df_out["Cumulative Length"] - df_out["Cumulative Length"].iloc[0]

        # Convertir la columna 'Time' a datetime
        df_out["Time_Seconds"] = (df_out["Time"] - df_out["Time"].iloc[0]).dt.total_seconds()

        # Aplicar demora para afinar los inicios de los tiempos
        df_out["Time_Seconds"] += displacement_out + last_offset #NOTE: Por si paso el medio de la intersección unos segundos después

    if not inbound:
        df_in = None

    if not outbound:
        df_out = None

    return df_in, df_out

def read_program(excel_path: str) -> tuple[dict, list, list]:
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active #NOTE: Podría generar un error

    n = 0 #NOTE: Uso para saltos de filas
    programs = {}
    offsets = []
    distances = [0]
    speeds = {
        "inbound": [],
        "outbound": [],
    }
    while True:
        name = ws.cell(row=3+2*n, column=1).value
        if name == None:
            break

        # Guardar más información
        id = n+1
        red =               ws.cell(row=3+2*n, column=12).value

        inbound_green =     ws.cell(row=3+2*n, column=6).value
        inbound_blue =      ws.cell(row=3+2*n, column=7).value
        inbound_leadlag =   ws.cell(row=3+2*n, column=8).value

        outbound_green =    ws.cell(row=3+2*n, column=9).value
        outbound_blue =     ws.cell(row=3+2*n, column=10).value
        outbound_leadlag =  ws.cell(row=3+2*n, column=11).value

        # Construcción del programa dinámico
        program = {
            "inbound": _get_color_values(inbound_leadlag, inbound_green, inbound_blue) + [["yellow", 3], ["red", red]],
            "outbound": _get_color_values(outbound_leadlag, outbound_green, outbound_blue) + [["yellow", 3], ["red", red]],
        }
        
        # Desfases
        offset = ws.cell(row=3+2*n, column=5).value
        offsets.append(offset)

        # Distancias
        distance = ws.cell(row=4+2*n, column=2).value
        if distance != None:
            distances.append(distance)

        # Velocidades
        in_speed = ws.cell(row=4+2*n, column=3).value
        if in_speed != None:
            in_speed_ms = in_speed / 3.6
            speeds["inbound"].append(in_speed_ms)
        out_speed = ws.cell(row=4+2*n, column=4).value
        if out_speed != None:
            out_speed_ms = out_speed / 3.6
            speeds["outbound"].append(out_speed_ms)

        # Programación
        programs[id] = program
        n += 1
    
    return programs, offsets, distances, speeds

def start_algorithm(original_programs: dict, offsets: list, distances: list, speeds: dict, df_in = None, df_out = None) -> None:
    # Configuración de las intersecciones
    intersection_positions = [sum(distances)-sum(distances[:i+1]) for i in range(len(distances))]
    # offsets = [0, 2, 28, 38] # Dulanto 29-01-25

    programs = deepcopy(original_programs)
    # Aplicando desfases
    for i, (intersection, dict_bounds) in enumerate(original_programs.items()):
        for bound, program_list in dict_bounds.items():
            new_program = apply_offset(program_list, offsets[i])
            programs[intersection][bound] = new_program

    light_cycles = {number: sum([value for color_key, value in dict_bounds["inbound"] if color_key != "yellow"]) for number, dict_bounds in programs.items()}
    time_max = max(light_cycles.values())*3  # Duración máxima del gráfico (en segundos)


    fig, ax = plt.subplots(figsize=(10, 6))
    ax.set_xlim(0, time_max)
    ax.set_xlabel("Time (s)")
    ax.set_ylabel("Distance (m)")
    ax.set_title("Time-Space Diagram")

    draw_band(ax, intersection_positions, programs, speeds, offsets, time_max, light_cycles, direction="inbound", color="green")
    draw_band(ax, intersection_positions, programs, speeds, offsets, time_max, light_cycles, direction="outbound", color="blue")

    # Graficar datos de GPS
    if df_in is not None:
        ax.plot(df_in["Time_Seconds"], df_in["Distance"], marker='o', linestyle='-', color='purple', label="Vehicle In Tracking")

    if df_out is not None:
        ax.plot(df_out["Time_Seconds"], df_out["Distance"], marker='o', linestyle='-', color='cyan', label="Vehicle Out Tracking")

    ax.legend()

    # Dibujar semáforos
    for (i, position), program, cycle in zip(enumerate(intersection_positions), programs.values(), light_cycles.values()):
        draw_light(ax, position, cycle, time_max, program)

    # Ajustes finales
    ax.grid(True, linestyle="--", alpha=0.7)
    plt.tight_layout()
    plt.show()